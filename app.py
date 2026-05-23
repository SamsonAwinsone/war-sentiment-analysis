"""
=============================================================================
app.py  —  War Sentiment Analysis  |  Flask Backend  (v3)
=============================================================================
Live data: BBC, Al Jazeera, Reuters, RT News, Google News (RSS + scraping),
           Reddit PRAW (optional), YouTube Data API v3 (optional).

Endpoints:
  GET /              → dashboard UI
  GET /fetch-data    → collect fresh data + run NLP pipeline
  GET /results       → full analysis JSON (charts/topics)
  GET /posts         → flat list with filters (source/sentiment/category/q)
  GET /articles      → articles grouped with comments nested inside
  GET /export-excel  → download Excel workbook (4 sheets)
=============================================================================
"""

import os, re, time, json, logging, hashlib
from datetime import datetime, timezone
from collections import defaultdict, deque
from pathlib import Path
from io import BytesIO
from threading import Thread
from concurrent.futures import ThreadPoolExecutor, as_completed

import feedparser
import requests
from bs4 import BeautifulSoup
import nltk
import pandas as pd
from flask import Flask, jsonify, render_template, send_file, request, Response
from flask_cors import CORS
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── NLTK bootstrap ────────────────────────────────────────────────────────────
for _r in [("corpora/stopwords","stopwords"),
           ("tokenizers/punkt","punkt"),
           ("tokenizers/punkt_tab","punkt_tab")]:
    try:
        nltk.data.find(_r[0])
    except Exception:
        # Handle LookupError and OSError (path issues on some systems)
        try:
            nltk.download(_r[1], quiet=True)
        except Exception:
            pass

_BUNDLED_STOPS = {
    "i","me","my","myself","we","our","ours","ourselves","you","your","yours",
    "yourself","yourselves","he","him","his","himself","she","her","hers",
    "herself","it","its","itself","they","them","their","theirs","themselves",
    "what","which","who","whom","this","that","these","those","am","is","are",
    "was","were","be","been","being","have","has","had","having","do","does",
    "did","doing","a","an","the","and","but","if","or","because","as","until",
    "while","of","at","by","for","with","about","against","between","into",
    "through","during","before","after","above","below","to","from","up","down",
    "in","out","on","off","over","under","again","further","then","once","here",
    "there","when","where","why","how","all","both","each","few","more","most",
    "other","some","such","no","nor","not","only","own","same","so","than","too",
    "very","s","t","can","will","just","don","should","now","d","ll","m","o",
    "re","ve","y","ain","aren","couldn","didn","doesn","hadn","hasn","haven",
    "isn","ma","mightn","mustn","needn","shan","shouldn","wasn","weren","won",
    "wouldn","said","say","says","also","may","get","go","make","take","come",
    "see","know","would","could","like","one","two","three","even","still",
    "reuters","bbc","aljazeera","monday","tuesday","wednesday","thursday",
    "friday","saturday","sunday","january","february","march","april","june",
    "july","august","september","october","november","december","new","year",
    "time","us","article","report","news","told","according","official",
}

try:
    from nltk.corpus import stopwords as _nltk_sw
    _NLTK_STOPS = set(_nltk_sw.words("english"))
except Exception:
    _NLTK_STOPS = set()

try:
    from nltk.tokenize import word_tokenize
except Exception:
    def word_tokenize(text):
        return re.sub(r"[^a-zA-Z\s]", " ", text).split()

try:
    from gensim import corpora
    from gensim.models import LdaModel
    GENSIM_OK = True
except ImportError:
    GENSIM_OK = False

# =============================================================================
# CONFIG
# =============================================================================

REDDIT_CLIENT_ID     = os.environ.get("REDDIT_CLIENT_ID",     "")
REDDIT_CLIENT_SECRET = os.environ.get("REDDIT_CLIENT_SECRET", "")
REDDIT_USER_AGENT    = "war_sentiment_bot/1.0"
YOUTUBE_API_KEY      = os.environ.get("YOUTUBE_API_KEY",      "")

KEYWORDS = ["iran", "israel", "hezbollah", "hormuz", "gaza",
            "tehran", "missile", "drone", "nuclear", "war",
            "middle east", "hamas", "netanyahu", "khamenei",
            "idf", "irgc", "beirut", "sanctions"]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# Global request/session settings
REQUEST_TIMEOUT = int(os.environ.get("REQUEST_TIMEOUT", "8"))  # seconds per request
session = requests.Session()
session.headers.update(HEADERS)

DATA_DIR   = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
CACHE_FILE = DATA_DIR / "cache.json"

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

app  = Flask(__name__)
CORS(app)
# Asset version for cache-busting static files (updated on server start)
app.jinja_env.globals['ASSET_VER'] = str(int(time.time()))

# =============================================================================
# UTILITIES
# =============================================================================

def _kw_match(text: str) -> bool:
    t = text.lower()
    return any(k in t for k in KEYWORDS)

def _clean_html(raw: str) -> str:
    return BeautifulSoup(raw or "", "html.parser").get_text(separator=" ")

def _save_cache(records: list):
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump({"timestamp": datetime.now(timezone.utc).isoformat(),
                       "records": records}, f)
    except Exception as e:
        log.warning(f"[Cache] save failed: {e}")

def _load_cache() -> list:
    try:
        if CACHE_FILE.exists():
            with open(CACHE_FILE, encoding="utf-8") as f:
                data = json.load(f)
                records = data.get("records", [])
                log.info(f"[Cache] loaded {len(records)} records")
                return records
    except json.JSONDecodeError as e:
        log.warning(f"[Cache] load failed: {e}. Renaming corrupt cache file")
        try:
            corrupt_path = CACHE_FILE.parent / f"cache.json.corrupt.{int(time.time())}"
            CACHE_FILE.replace(corrupt_path)
            log.info(f"[Cache] corrupt cache renamed to {corrupt_path.name}")
        except Exception as ex:
            log.warning(f"[Cache] failed to rename corrupt cache: {ex}")
    except Exception as e:
        log.warning(f"[Cache] load failed: {e}")
    return []

# =============================================================================
# SECTION A — LIVE DATA COLLECTION
# =============================================================================

RSS_SOURCES = [
    ("BBC",         "https://feeds.bbci.co.uk/news/world/middle_east/rss.xml"),
    ("BBC",         "https://feeds.bbci.co.uk/news/world/rss.xml"),
    ("Al Jazeera",  "https://www.aljazeera.com/xml/rss/all.xml"),
    ("Reuters",     "https://feeds.reuters.com/reuters/worldNews"),
    ("Reuters",     "https://feeds.reuters.com/Reuters/worldnews"),
    ("RT News",     "https://www.rt.com/rss/news/"),
    ("Google News", "https://news.google.com/rss/search?q=Iran+Israel+war&hl=en-US&gl=US&ceid=US:en"),
    ("Google News", "https://news.google.com/rss/search?q=Iran+US+military+strike&hl=en-US&gl=US&ceid=US:en"),
    ("Google News", "https://news.google.com/rss/search?q=Israel+Iran+nuclear&hl=en-US&gl=US&ceid=US:en"),
    ("Google News", "https://news.google.com/rss/search?q=Hezbollah+Iran+Israel&hl=en-US&gl=US&ceid=US:en"),
    ("Sky News",    "https://feeds.skynews.com/feeds/rss/world.xml"),
    ("Guardian",    "https://www.theguardian.com/world/middleeast/rss"),
    ("DW",          "https://rss.dw.com/rdf/rss-en-all"),
    ("France 24",   "https://www.france24.com/en/middle-east/rss"),
]

SCRAPE_TARGETS = [
    ("BBC",        "https://www.bbc.com/news/world/middle_east"),
    ("Al Jazeera", "https://www.aljazeera.com/where/iran/"),
    ("Al Jazeera", "https://www.aljazeera.com/where/israel/"),
    ("Reuters",    "https://www.reuters.com/world/middle-east/"),
    ("RT News",    "https://www.rt.com/news/middle-east/"),
]


def fetch_rss() -> list:
    def _fetch_rss_source(outlet: str, url: str) -> list:
        local = []
        try:
            log.info(f"[RSS] Fetching from {outlet}: {url}")
            try:
                resp = session.get(url, timeout=REQUEST_TIMEOUT)
                resp.raise_for_status()
                feed = feedparser.parse(resp.content)
            except Exception as e:
                log.warning(f"[RSS] {outlet}: request failed: {e}")
                return []

            if not hasattr(feed, 'entries') or not feed.entries:
                log.warning(f"[RSS] {outlet}: No entries found")
                return []

            seen = set()
            count = 0
            for entry in feed.entries:
                title = entry.get("title", "")
                summary = _clean_html(entry.get("summary", entry.get("description", "")))
                link = entry.get("link", "")
                pub = entry.get("published",
                      entry.get("updated",
                      datetime.now(timezone.utc).isoformat()))
                text = f"{title}. {summary}"
                if not text or len(text) < 20 or not _kw_match(text):
                    continue
                uid = hashlib.md5(link.encode()).hexdigest() if link else hashlib.md5(text.encode()).hexdigest()
                if uid in seen:
                    continue
                seen.add(uid)
                local.append({
                    "source": outlet,
                    "title": title[:200],
                    "text": text[:1200],
                    "url": link,
                    "date": pub,
                    "method": "RSS Feed",
                    "type": "article",
                })
                count += 1
                if count >= 15:
                    break

            if count:
                log.info(f"[RSS] {outlet}: +{count} articles")
            else:
                log.warning(f"[RSS] {outlet}: No matching articles found")
        except Exception as e:
            log.warning(f"[RSS] {outlet} failed: {e}")
        return local

    records = []
    with ThreadPoolExecutor(max_workers=6) as ex:
        futures = {ex.submit(_fetch_rss_source, outlet, url): (outlet, url)
                   for outlet, url in RSS_SOURCES}
        for fut in as_completed(futures):
            try:
                records.extend(fut.result() or [])
            except Exception as e:
                outlet, url = futures.get(fut, (None, None))
                log.warning(f"[RSS] {outlet or url} failed in executor: {e}")

    log.info(f"[RSS] Total: {len(records)} articles collected")
    return records


def scrape_headlines() -> list:
    def _fetch_scrape_target(outlet: str, url: str) -> list:
        local = []
        try:
            log.info(f"[Scrape] Fetching from {outlet}: {url}")
            r = session.get(url, timeout=max(4, REQUEST_TIMEOUT))
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            cnt = 0
            for tag in soup.find_all(["h1", "h2", "h3", "h4"]):
                a = tag.find("a", href=True)
                if not a:
                    continue
                title = a.get_text(separator=" ").strip()
                href = a["href"]
                if len(title) < 15 or not _kw_match(title):
                    continue
                if href.startswith("/"):
                    from urllib.parse import urlparse
                    base = urlparse(url)
                    href = f"{base.scheme}://{base.netloc}{href}"
                local.append({
                    "source": outlet,
                    "title": title[:200],
                    "text": title,
                    "url": href,
                    "date": datetime.now(timezone.utc).isoformat(),
                    "method": "Web Scraping",
                    "type": "article",
                })
                cnt += 1
                if cnt >= 10:
                    break
            if cnt:
                log.info(f"[Scrape] {outlet}: +{cnt} articles")
            else:
                log.warning(f"[Scrape] {outlet}: No matching headlines found")
        except Exception as e:
            log.warning(f"[Scrape] {outlet} failed: {e}")
        return local

    records = []
    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = {ex.submit(_fetch_scrape_target, outlet, url): (outlet, url)
                   for outlet, url in SCRAPE_TARGETS}
        for fut in as_completed(futures):
            try:
                records.extend(fut.result() or [])
            except Exception as e:
                outlet, url = futures.get(fut, (None, None))
                log.warning(f"[Scrape] {outlet or url} failed in executor: {e}")

    log.info(f"[Scrape] Total: {len(records)} articles collected")
    return records


def fetch_reddit() -> list:
    if not REDDIT_CLIENT_ID or not REDDIT_CLIENT_SECRET:
        return []
    try:
        import praw
        reddit  = praw.Reddit(
            client_id=REDDIT_CLIENT_ID, client_secret=REDDIT_CLIENT_SECRET,
            user_agent=REDDIT_USER_AGENT)
        records = []
        subs    = ["worldnews","geopolitics","CredibleDefense","iran","IsraelPalestine"]
        kws     = ["Iran Israel war","Iran missile attack","Iran nuclear Israel"]
        for sub in subs:
            for kw in kws:
                for post in reddit.subreddit(sub).search(
                        kw, sort="top", time_filter="year", limit=20):
                    post_url = f"https://reddit.com{post.permalink}"
                    records.append({
                        "source": f"Reddit/r/{sub}", "title": post.title,
                        "text":   f"{post.title}. {post.selftext[:600]}",
                        "url":    post_url,
                        "date":   datetime.fromtimestamp(
                                  post.created_utc, tz=timezone.utc).isoformat(),
                        "method": "Reddit API (PRAW)", "type": "article",
                    })
                    post.comments.replace_more(limit=0)
                    for c in list(post.comments)[:20]:
                        if len(c.body) > 30:
                            records.append({
                                "source":     f"Reddit/r/{sub}",
                                "title":      post.title[:100],
                                "text":       c.body[:600],
                                "url":        f"https://reddit.com{post.permalink}#{c.id}",
                                "date":       datetime.fromtimestamp(
                                              c.created_utc, tz=timezone.utc).isoformat(),
                                "method":     "Reddit API (PRAW)",
                                "type":       "comment",
                                "parent_url": post_url,
                            })
                # small polite pause
                time.sleep(0.2)
        log.info(f"[Reddit] {len(records)} records")
        return records
    except Exception as e:
        log.warning(f"[Reddit] {e}")
        return []


def fetch_youtube() -> list:
    if not YOUTUBE_API_KEY:
        return []
    base, records = "https://www.googleapis.com/youtube/v3", []
    channels = {
        "BBC News":   "UCnUYZLuoy1rq1aVMwx4aTzw",
        "Al Jazeera": "UCNye-wNBqNL5ZzHSJdba7Xg",
    }
    for ch_name, ch_id in channels.items():
        try:
            r = session.get(f"{base}/search", params={
                "key": YOUTUBE_API_KEY, "channelId": ch_id,
                "q": "Iran Israel war", "part": "snippet",
                "type": "video", "maxResults": 8, "order": "relevance",
            }, timeout=max(6, REQUEST_TIMEOUT))
            r.raise_for_status()
            for v in r.json().get("items", []):
                vid       = v["id"]["videoId"]
                vtitle    = v["snippet"]["title"]
                pub       = v["snippet"]["publishedAt"]
                video_url = f"https://youtube.com/watch?v={vid}"
                if not _kw_match(vtitle):
                    continue
                records.append({
                    "source": f"YouTube/{ch_name}", "title": vtitle[:120],
                    "text":   vtitle, "url":   video_url,
                    "date":   pub,    "method": "YouTube Data API v3",
                    "type":   "article",
                })
                # Limit comment fetch to reduce latency
                cr = session.get(f"{base}/commentThreads", params={
                    "key": YOUTUBE_API_KEY, "videoId": vid,
                    "part": "snippet", "maxResults": 20, "order": "relevance",
                }, timeout=max(6, REQUEST_TIMEOUT))
                cr.raise_for_status()
                for item in cr.json().get("items", []):
                    sn = item["snippet"]["topLevelComment"]["snippet"]
                    records.append({
                        "source":     f"YouTube/{ch_name}",
                        "title":      vtitle[:120],
                        "text":       sn.get("textDisplay", ""),
                        "url":        f"{video_url}&lc={item['id']}",
                        "date":       sn.get("publishedAt", pub),
                        "method":     "YouTube Data API v3",
                        "type":       "comment",
                        "parent_url": video_url,
                    })
            # small polite pause between channel queries
            time.sleep(0.3)
        except Exception as e:
            log.warning(f"[YouTube] {ch_name}: {e}")
    return records


def collect_all_live() -> list:
    all_records = []
    log.info("=== Starting parallel live collection ===")

    # Run primary collectors in parallel to reduce wall-clock time
    collectors = [fetch_rss, scrape_headlines, fetch_reddit, fetch_youtube]
    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = {ex.submit(fn): fn.__name__ for fn in collectors}
        for fut in as_completed(futures):
            name = futures.get(fut, "collector")
            try:
                result = fut.result()
                if result:
                    all_records.extend(result)
                    log.info(f"[{name}] done: {len(result)} records")
                else:
                    log.warning(f"[{name}] returned no records")
            except Exception as e:
                log.warning(f"[{name}] failed: {e}")

    # Deduplicate articles by URL (keep all comments)
    seen_urls, unique = set(), []
    for r in all_records:
        key = r.get("url", "")
        if r.get("type", "article") == "article":
            if key and key in seen_urls:
                continue
            if key:
                seen_urls.add(key)
        unique.append(r)

    if len(unique) > 0:
        _save_cache(unique)
        log.info(f"=== Done: {len(unique)} records ===")
        return unique

    cached = _load_cache()
    return cached if cached else []


# =============================================================================
# SECTION B — PREPROCESSING
# =============================================================================

STOP_WORDS = _BUNDLED_STOPS | _NLTK_STOPS
URL_RE     = re.compile(r"https?://\S+|www\.\S+")
HTML_RE    = re.compile(r"<[^>]+>")
PUNCT_RE   = re.compile(r"[^a-zA-Z\s]")
SPACE_RE   = re.compile(r"\s+")


def preprocess(text: str) -> list:
    """Full NLP pipeline → list of clean tokens for LDA."""
    if not text or not isinstance(text, str):
        return []
    text = text.lower()
    text = HTML_RE.sub(" ", text)
    text = URL_RE.sub(" ",  text)
    text = PUNCT_RE.sub(" ", text)
    text = SPACE_RE.sub(" ", text).strip()
    try:
        tokens = word_tokenize(text)
    except Exception:
        tokens = text.split()
    return [t for t in tokens if t not in STOP_WORDS and len(t) >= 3]


def preprocess_for_vader(text: str) -> str:
    """Light clean preserving punctuation (VADER uses it for emphasis)."""
    text = HTML_RE.sub(" ", text or "")
    text = URL_RE.sub("", text)
    return SPACE_RE.sub(" ", text).strip()[:800]


# =============================================================================
# SECTION C — TOPIC MODELING (LDA via gensim)
# =============================================================================

TOPIC_LABELS = {
    0: "Military Operations & Strikes",
    1: "Geopolitical Tensions",
    2: "Nuclear Programme",
    3: "Economic Impact & Energy",
    4: "Humanitarian Crisis",
    5: "Media Narratives & Propaganda",
    6: "Diplomatic Relations",
    7: "Support / Opposition to War",
}


def run_lda(docs: list, n_topics: int = 7) -> list:
    if not GENSIM_OK or len(docs) < 10:
        return _fallback_topics(docs)
    try:
        dictionary = corpora.Dictionary(docs)
        dictionary.filter_extremes(no_below=2, no_above=0.85)
        corpus = [dictionary.doc2bow(d) for d in docs]
        lda = LdaModel(corpus=corpus, id2word=dictionary,
                       num_topics=n_topics, random_state=42,
                       passes=10, alpha="auto", per_word_topics=False)
        topics = []
        for i in range(n_topics):
            kw_pairs = lda.show_topic(i, topn=10)
            keywords = [w for w, _ in kw_pairs]
            weight   = float(sum(p for _, p in kw_pairs))
            topics.append({
                "id": i, "label": TOPIC_LABELS.get(i, f"Topic {i+1}"),
                "keywords": keywords, "weight": round(weight * 100, 2),
            })
        total = sum(t["weight"] for t in topics) or 1
        for t in topics:
            t["weight"] = round(t["weight"] / total * 100, 1)
        return topics
    except Exception as e:
        log.warning(f"[LDA] {e}")
        return _fallback_topics(docs)


def _fallback_topics(docs: list) -> list:
    seeds = [
        ("Military Operations & Strikes", ["missile","drone","strike","attack","military","bomb","idf","irgc","ballistic","intercept"]),
        ("Geopolitical Tensions",         ["tension","sanction","threat","ally","alliance","nato","russia","china","region","diplomacy"]),
        ("Nuclear Programme",             ["nuclear","uranium","enrich","warhead","iaea","facility","weapon","natanz","proliferation","inspectors"]),
        ("Economic Impact & Energy",      ["oil","price","barrel","hormuz","energy","supply","tanker","economic","market","opec"]),
        ("Humanitarian Crisis",           ["civilian","hospital","aid","refugee","death","kill","ceasefire","humanitarian","suffer","child"]),
        ("Media Narratives & Propaganda", ["media","propaganda","fake","bias","narrative","coverage","disinformation","misinformation","report","claim"]),
        ("Diplomatic Relations",          ["diplomacy","talks","negotiation","ceasefire","agreement","minister","foreign","ambassador","un","resolution"]),
        ("Support / Opposition to War",   ["protest","support","oppose","rally","march","condemn","war","peace","opposition","government"]),
    ]
    flat = [w for doc in docs for w in doc]
    freq = defaultdict(int)
    for w in flat:
        freq[w] += 1
    results = []
    topic_scores = []
    for i, (label, kws) in enumerate(seeds):
        # score topic by frequency of its seed keywords in the corpus
        score = 0
        kws_found = []
        for kw in kws:
            cnt = freq.get(kw, 0)
            if cnt > 0:
                kws_found.append(kw)
            score += cnt
        topic_scores.append((i, label, score, kws_found))

    # compute weights proportionally (avoid divide by zero)
    total_score = sum(t[2] for t in topic_scores) or 1
    for i, label, score, kws_found in topic_scores:
        # pick top keywords for display (most frequent tokens overlapping seeds)
        kw_candidates = {w: freq.get(w, 0) for w in kws_found}
        # include top corpus tokens that share a prefix with seeds
        for w, cnt in freq.items():
            for s in seeds[i][1]:
                if s[:4] in w and w not in kw_candidates:
                    kw_candidates[w] = kw_candidates.get(w, 0) + cnt
        top = sorted(kw_candidates, key=kw_candidates.get, reverse=True)[:10]
        results.append({
            "id": i,
            "label": label,
            "keywords": top,
            "weight": round(score / total_score * 100, 1)
        })
    return results


# =============================================================================
# SECTION D — SENTIMENT ANALYSIS (VADER)
# =============================================================================

vader = SentimentIntensityAnalyzer()

THEMATIC_KWS = {
    "Military Operations": [
        "missile","drone","strike","attack","bomb","idf","irgc","military",
        "soldier","war","combat","ballistic","intercept","airstrike","navy",
        "army","artillery","rocket","weapon","explosion","kill","casualty",
    ],
    "Geopolitical Tensions": [
        "tension","sanction","threat","alliance","nato","russia","china",
        "regime","power","sovereignty","territory","proxy","conflict",
        "escalate","deterrence","nuclear","hezbollah","hamas","iran","israel",
    ],
    "Economic Impact": [
        "oil","barrel","price","hormuz","energy","supply","tanker",
        "economy","market","export","opec","inflation","commodity",
        "gdp","trade","sanction","cost","fuel","gas","shipping",
    ],
    "Media Narratives": [
        "media","propaganda","fake","bias","narrative","coverage","report",
        "disinformation","misinformation","claim","framing","censor",
        "journalist","news","outlet","western","bbc","aljazeera","rt",
    ],
    "Support for War": [
        "support","oppose","protest","rally","march","condemn","peace",
        "ally","government","policy","public","opinion","debate",
        "election","vote","international","solidarity","civilian",
    ],
}


def classify_article(text: str) -> str:
    lower  = text.lower()
    scores = {cat: sum(lower.count(kw) for kw in kws)
              for cat, kws in THEMATIC_KWS.items()}
    best   = max(scores, key=scores.get)
    return best if scores[best] > 0 else "General"


def analyse_sentiment(records: list) -> dict:
    global _enriched_records
    if not records:
        _enriched_records = []
        return {}

    rows = []
    for i, rec in enumerate(records):
        clean = preprocess_for_vader(rec.get("text", ""))
        if not clean:
            continue
        sc    = vader.polarity_scores(clean)
        comp  = sc["compound"]
        label = ("Positive" if comp >= 0.05
                 else "Negative" if comp <= -0.05
                 else "Neutral")
        cat   = classify_article(clean)
        rows.append({
            **rec,
            "id":        hashlib.md5(f"{rec.get('url','')}{i}".encode()).hexdigest(),
            "compound":  comp,
            "pos":       sc["pos"],
            "neu":       sc["neu"],
            "neg":       sc["neg"],
            "sentiment": label,
            "category":  cat,
        })

    _enriched_records = rows
    df = pd.DataFrame(rows)

    # Per-platform
    platform_stats = []
    for src, grp in df.groupby("source"):
        n   = len(grp)
        avg = round(float(grp["compound"].mean()), 4)
        pos = int((grp["sentiment"] == "Positive").sum())
        neu = int((grp["sentiment"] == "Neutral").sum())
        neg = int((grp["sentiment"] == "Negative").sum())
        platform_stats.append({
            "platform": src, "count": n, "avg_compound": avg,
            "positive": pos, "neutral": neu, "negative": neg,
            "pos_pct": round(pos/n*100, 1),
            "neu_pct": round(neu/n*100, 1),
            "neg_pct": round(neg/n*100, 1),
        })
    platform_stats.sort(key=lambda x: x["count"], reverse=True)

    total = len(df)
    pos_t = int((df["sentiment"] == "Positive").sum())
    neu_t = int((df["sentiment"] == "Neutral").sum())
    neg_t = int((df["sentiment"] == "Negative").sum())

    # Per-category
    category_stats = []
    for cat, grp in df.groupby("category"):
        n   = len(grp)
        avg = round(float(grp["compound"].mean()), 4)
        category_stats.append({
            "category": cat, "count": n, "avg_compound": avg,
            "positive": int((grp["sentiment"] == "Positive").sum()),
            "neutral":  int((grp["sentiment"] == "Neutral").sum()),
            "negative": int((grp["sentiment"] == "Negative").sum()),
        })
    category_stats.sort(key=lambda x: x["count"], reverse=True)

    # Timeline
    def parse_date(d):
        d_str = str(d or "").strip()
        for fmt in ["%a, %d %b %Y %H:%M:%S %z", "%a, %d %b %Y %H:%M:%S",
                    "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S%z",
                    "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"]:
            try:
                return datetime.strptime(d_str[:25], fmt).strftime("%Y-%m-%d")
            except Exception:
                continue
        m = re.search(r"\d{4}-\d{2}-\d{2}", d_str)
        return m.group(0) if m else None

    df["day"] = df["date"].apply(lambda d: parse_date(str(d)))
    timeline  = (df.dropna(subset=["day"]).groupby("day")["compound"]
                   .mean().round(4).reset_index()
                   .rename(columns={"compound": "avg_sentiment"})
                   .sort_values("day").to_dict("records"))
    result = {
        "total_articles": total,
        "platform_stats": platform_stats,
        "overall": {
            "positive": pos_t, "neutral": neu_t, "negative": neg_t,
            "pos_pct":  round(pos_t/total*100,1) if total else 0,
            "neu_pct":  round(neu_t/total*100,1) if total else 0,
            "neg_pct":  round(neg_t/total*100,1) if total else 0,
        },
        "category_stats": category_stats,
        "timeline":       timeline,
    }

    log.info(f"[Analysis] Generated analysis: total={total}, pos={pos_t}, neu={neu_t}, neg={neg_t}")
    return result


def build_collection_table(records: list) -> list:
    info = defaultdict(lambda: {"count":0,"methods":set(),"earliest":"","latest":""})
    for r in records:
        src = r["source"]
        info[src]["count"] += 1
        info[src]["methods"].add(r.get("method", "unknown"))
        d = str(r.get("date", ""))[:10]
        if d:
            if not info[src]["earliest"] or d < info[src]["earliest"]:
                info[src]["earliest"] = d
            if not info[src]["latest"] or d > info[src]["latest"]:
                info[src]["latest"] = d
    return [{"outlet": src, "count": v["count"],
              "earliest": v["earliest"] or "N/A",
              "latest":   v["latest"]   or "N/A",
              "methods":  ", ".join(v["methods"]),
              "keywords": "Iran, Israel, war, Hezbollah, Hormuz, nuclear"}
            for src, v in sorted(info.items(),
                                 key=lambda x: x[1]["count"], reverse=True)]


# =============================================================================
# IN-MEMORY STATE
# =============================================================================

_live_records     = []
_analysis         = {}
_topics           = []
_table            = []
_last_updated     = None
_is_cached        = False
_enriched_records = []
_bg_fetching      = False
_bg_last_start    = None
_bg_last_end      = None
_bg_last_count    = 0
_recent_logs      = deque(maxlen=400)


class BufferLogHandler(logging.Handler):
    def emit(self, record):
        try:
            msg = self.format(record)
            # keep recent formatted log lines
            _recent_logs.append(msg)
        except Exception:
            pass

# Attach buffer handler to capture recent logs for UI
_buf_handler = BufferLogHandler()
_buf_handler.setLevel(logging.INFO)
_buf_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logging.getLogger().addHandler(_buf_handler)


# Initialize with cached data on startup
def _init_with_cache():
    global _live_records, _table, _last_updated, _is_cached, _analysis
    try:
        cached = _load_cache()
        if cached:
            _live_records = cached
            _table = build_collection_table(_live_records)
            _last_updated = datetime.now(timezone.utc).isoformat()
            _run_full_analysis()
            _is_cached = True
            log.info(f"[INIT] Loaded {len(cached)} cached records on startup")
        else:
            log.warning("[INIT] No cache found on startup")
    except Exception as e:
        log.error(f"[INIT] Failed to load cache: {e}")


def _run_full_analysis():
    global _analysis, _topics
    try:
        docs     = [preprocess(r.get("text","")) for r in _live_records]
        docs     = [d for d in docs if len(d) >= 5]
        _topics   = run_lda(docs, n_topics=7) if len(docs) >= 10 else _fallback_topics(docs)
        _analysis = analyse_sentiment(_live_records)
    except Exception as e:
        log.error(f"[Analysis] {e}", exc_info=True)


def _background_fetch_live_data():
    """Background thread to fetch live data without blocking HTTP response"""
    global _live_records, _table, _last_updated, _is_cached
    global _bg_fetching, _bg_last_start, _bg_last_end, _bg_last_count
    try:
        _bg_fetching = True
        _bg_last_start = datetime.now(timezone.utc).isoformat()
        _bg_last_end = None
        _bg_last_count = 0
        log.info("[BG] Starting background live data collection...")

        collected = collect_all_live()

        if collected:
            _live_records = collected
            _is_cached = False
            _bg_last_count = len(collected)
            log.info(f"[BG] SUCCESS: Collected {_bg_last_count} LIVE records")
        else:
            log.warning("[BG] Live collection empty, keeping cache")
            _is_cached = True

        _table = build_collection_table(_live_records)
        _last_updated = datetime.now(timezone.utc).isoformat()
        _run_full_analysis()
        _bg_last_end = datetime.now(timezone.utc).isoformat()
        log.info("[BG] Analysis complete!")

    except Exception as e:
        log.error(f"[BG] Background fetch failed: {e}", exc_info=True)
        _is_cached = True
        _bg_last_end = datetime.now(timezone.utc).isoformat()
    finally:
        _bg_fetching = False


# =============================================================================
# FLASK ROUTES
# =============================================================================
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/fetch-status")
def fetch_status():
    """Return background fetch status for frontend polling"""
    try:
        return jsonify({
            "bg_fetching": bool(_bg_fetching),
            "last_start": _bg_last_start,
            "last_end": _bg_last_end,
            "last_count": int(_bg_last_count),
            "last_updated": _last_updated,
            "logs": list(_recent_logs)[-60:],
        })
    except Exception as e:
        log.error(f"[/fetch-status] {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500
@app.route("/fetch-data")
def fetch_data():
    global _live_records, _table, _last_updated, _is_cached
    try:
        # Start background thread to fetch live data
        bg_thread = Thread(target=_background_fetch_live_data, daemon=True)
        bg_thread.start()
        
        # Load cache immediately for instant response
        cached = _load_cache()
        if cached:
            _live_records = cached
            _table = build_collection_table(_live_records)
            _last_updated = datetime.now(timezone.utc).isoformat()
            _run_full_analysis()
            _is_cached = True
            
            return jsonify({
                "status": "fetching_live",
                "message": "Fetching live data in background... showing cached data",
                "total_fetched": len(_live_records),
                "sources": len(_table),
                "is_cached": True,
                "last_updated": _last_updated,
                "collection_table": _table,
            })
        elif _live_records:
            _table = build_collection_table(_live_records)
            _last_updated = datetime.now(timezone.utc).isoformat()
            _run_full_analysis()
            return jsonify({
                "status": "fetching_live",
                "message": "Fetching live data in background... showing current data",
                "total_fetched": len(_live_records),
                "sources": len(_table),
                "is_cached": bool(_is_cached),
                "last_updated": _last_updated,
                "collection_table": _table,
            })
        else:
            # No cache, just return empty with message
            return jsonify({
                "status": "fetching_live",
                "message": "Fetching live data (no cache available yet)",
                "total_fetched": 0,
                "sources": 0,
                "is_cached": False,
                "last_updated": None,
                "collection_table": [],
            })
            
    except Exception as e:
        log.error(f"[/fetch-data] {e}", exc_info=True)
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route('/__info__')
def _internal_info():
    try:
        return jsonify({
            'root_path': app.root_path,
            'static_folder': app.static_folder,
            'template_folder': app.template_folder,
            'asset_ver': app.jinja_env.globals.get('ASSET_VER')
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/asset/script.js')
def serve_script_nocache():
    """Serve the main script with no-cache headers to avoid browser caching during dev."""
    try:
        path = Path(app.root_path) / 'static' / 'js' / 'script.js'
        resp = send_file(str(path), mimetype='application/javascript')
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp
    except Exception as e:
        log.error(f"[/asset/script.js] {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/events')
def stream_events():
    """Server-Sent Events stream for background fetch status updates."""
    def gen():
        last = None
        try:
            while True:
                payload = json.dumps({
                    "bg_fetching": bool(_bg_fetching),
                    "last_start": _bg_last_start,
                    "last_end": _bg_last_end,
                    "last_count": int(_bg_last_count),
                    "last_updated": _last_updated,
                    "logs": list(_recent_logs)[-200:],
                })
                if payload != last:
                    yield f"data: {payload}\n\n"
                    last = payload
                time.sleep(1.5)
        except GeneratorExit:
            return
    return Response(gen(), mimetype='text/event-stream')


@app.route("/results")
def results():
    try:
        if not _analysis:
            log.error("[/results] No analysis data available")
            return jsonify({"error": "No results yet. Call /fetch-data first."}), 400
        
        result = {
            "last_updated":     _last_updated,
            "is_cached":        _is_cached,
            "collection_table": _table,
            "sentiment":        _analysis,
            "topics":           _topics,
        }
        
        log.info(f"[/results] Returning analysis - Articles: {_analysis.get('total_articles', 0)}, "
                f"Positive: {_analysis.get('overall', {}).get('positive', 0)}, "
                f"Topics: {len(_topics)}")
        
        return jsonify(result)
    except Exception as e:
        log.error(f"[/results] Exception: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/posts")
def get_posts():
    """Flat filtered list — used by Posts table."""
    if not _enriched_records:
        return jsonify([])
    source_f    = request.args.get("source",    "").strip()
    sentiment_f = request.args.get("sentiment", "").strip()
    category_f  = request.args.get("category",  "").strip()
    q           = request.args.get("q",         "").strip().lower()
    limit       = int(request.args.get("limit", 500))

    res = _enriched_records
    if source_f    and source_f    != "all": res = [r for r in res if r.get("source")    == source_f]
    if sentiment_f and sentiment_f != "all": res = [r for r in res if r.get("sentiment") == sentiment_f]
    if category_f  and category_f  != "all": res = [r for r in res if r.get("category")  == category_f]
    if q: res = [r for r in res
                 if q in r.get("text","").lower() or q in r.get("title","").lower()]

    return jsonify([{
        "id":           r.get("id",""),
        "source":       r.get("source",""),
        "title":        r.get("title","")[:200],
        "text_preview": r.get("text","")[:300],
        "text_full":    r.get("text",""),
        "sentiment":    r.get("sentiment","Neutral"),
        "compound":     r.get("compound",0),
        "pos":          r.get("pos",0),
        "neu":          r.get("neu",0),
        "neg":          r.get("neg",0),
        "category":     r.get("category","General"),
        "date":         r.get("date",""),
        "url":          r.get("url",""),
        "method":       r.get("method",""),
        "type":         r.get("type","article"),
    } for r in res[:limit]])


@app.route("/articles")
def get_articles():
    """
    Articles grouped with their nested comments.
    Each article object includes a `comments` array of associated comment records.
    Comments are matched by parent_url == article url.
    """
    if not _enriched_records:
        return jsonify([])

    articles = [r for r in _enriched_records
                if r.get("type","article") == "article"]
    comments = [r for r in _enriched_records
                if r.get("type") == "comment"]

    # Index comments by parent_url
    by_parent = defaultdict(list)
    for c in comments:
        key = c.get("parent_url") or c.get("url","")
        by_parent[key].append(c)

    grouped = []
    for art in articles:
        art_url  = art.get("url","")
        art_cmts = sorted(by_parent.get(art_url, []),
                          key=lambda x: str(x.get("date","")))
        grouped.append({
            "id":            art.get("id",""),
            "source":        art.get("source",""),
            "title":         art.get("title","")[:200],
            "text":          art.get("text",""),
            "sentiment":     art.get("sentiment","Neutral"),
            "compound":      round(art.get("compound",0), 4),
            "pos":           art.get("pos",0),
            "neu":           art.get("neu",0),
            "neg":           art.get("neg",0),
            "category":      art.get("category","General"),
            "date":          art.get("date",""),
            "url":           art_url,
            "method":        art.get("method",""),
            "comment_count": len(art_cmts),
            "comments": [{
                "id":        c.get("id",""),
                "source":    c.get("source",""),
                "text":      c.get("text",""),
                "sentiment": c.get("sentiment","Neutral"),
                "compound":  round(c.get("compound",0), 4),
                "date":      c.get("date",""),
                "url":       c.get("url",""),
            } for c in art_cmts[:50]],
        })

    grouped.sort(key=lambda x: str(x.get("date","")), reverse=True)
    limit = int(request.args.get("limit", 200))
    return jsonify(grouped[:limit])


@app.route("/export-excel")
def export_excel():
    """Download full dataset as formatted 4-sheet Excel workbook."""
    if not _enriched_records:
        return jsonify({"error": "No data. Run /fetch-data first."}), 400
    if not OPENPYXL_OK:
        return jsonify({"error": "openpyxl not installed. pip install openpyxl"}), 500

    wb = Workbook()
    wb.remove(wb.active)

    HDR_FILL_DARK  = PatternFill("solid", fgColor="0F172A")
    HDR_FONT_WHITE = Font(bold=True, color="FFFFFF", size=11)
    SUB_FILL       = PatternFill("solid", fgColor="1E293B")
    SUB_FONT       = Font(bold=True, color="94A3B8")
    POS_FILL       = PatternFill("solid", fgColor="14532D")
    NEG_FILL       = PatternFill("solid", fgColor="7F1D1D")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16

    def srow(row_data, fill=None, bold=False):
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            if fill: cell.fill = fill
            if bold: cell.font = Font(bold=True, color=("FFFFFF" if fill else "E2E8F0"))
            cell.alignment = Alignment(vertical="center")

    srow(["WAR SENTIMENT ANALYSIS — US / Israel–Iran Conflict"],
         fill=HDR_FILL_DARK, bold=True)
    ws["A1"].font = Font(bold=True, color="60A5FA", size=13)
    ws.row_dimensions[1].height = 24
    srow([f"Generated: {_last_updated or 'N/A'}"])
    srow(["Total Records", len(_enriched_records)])
    srow(["Data Status",  "Cached" if _is_cached else "Live"])
    ws.append([])
    srow(["SENTIMENT DISTRIBUTION"], fill=SUB_FILL, bold=True)
    srow(["Sentiment","Count","Percentage"], fill=SUB_FILL)
    if _analysis.get("overall"):
        ov = _analysis["overall"]
        ws.append(["Positive", ov["positive"], f"{ov['pos_pct']}%"])
        ws.append(["Neutral",  ov["neutral"],  f"{ov['neu_pct']}%"])
        ws.append(["Negative", ov["negative"], f"{ov['neg_pct']}%"])
    ws.append([])
    srow(["PLATFORM BREAKDOWN"], fill=SUB_FILL, bold=True)
    srow(["Platform","Articles","Avg Score","Positive","Neutral","Negative","+%","=%","-%"],
         fill=SUB_FILL)
    for p in _analysis.get("platform_stats",[]):
        ws.append([p["platform"], p["count"], round(p["avg_compound"],3),
                   p["positive"], p["neutral"], p["negative"],
                   f"{p['pos_pct']}%", f"{p['neu_pct']}%", f"{p['neg_pct']}%"])
    ws.append([])
    srow(["THEMATIC BREAKDOWN"], fill=SUB_FILL, bold=True)
    srow(["Category","Articles","Avg Score","Positive","Neutral","Negative"],
         fill=SUB_FILL)
    for c in _analysis.get("category_stats",[]):
        ws.append([c["category"], c["count"], round(c["avg_compound"],3),
                   c["positive"], c["neutral"], c["negative"]])

    ARTICLE_HEADERS = ["ID","Source","Type","Title","Content","Sentiment",
                       "Score","Pos%","Neu%","Neg%","Category","Date","URL","Method"]
    COL_WIDTHS      = [10, 18, 10, 42, 52, 12, 10, 8, 8, 8, 24, 18, 42, 20]

    def make_sheet(name, rows, header_fill):
        sh = wb.create_sheet(name)
        sh.append(ARTICLE_HEADERS)
        for cell in sh[1]:
            cell.fill = header_fill
            cell.font = HDR_FONT_WHITE
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for r in rows:
            sh.append([
                r.get("id","")[:8],    r.get("source",""),
                r.get("type","article"), r.get("title","")[:120],
                r.get("text","")[:500], r.get("sentiment","Neutral"),
                round(r.get("compound",0),4),
                round(r.get("pos",0)*100,1),
                round(r.get("neu",0)*100,1),
                round(r.get("neg",0)*100,1),
                r.get("category","General"),
                str(r.get("date",""))[:19],
                r.get("url",""),        r.get("method",""),
            ])
        for i, w in enumerate(COL_WIDTHS, 1):
            sh.column_dimensions[chr(64+i)].width = w
        return sh

    make_sheet("All Articles",      _enriched_records, HDR_FILL_DARK)
    make_sheet("Positive Articles",
               [r for r in _enriched_records if r.get("sentiment") == "Positive"],
               POS_FILL)
    make_sheet("Negative Articles",
               [r for r in _enriched_records if r.get("sentiment") == "Negative"],
               NEG_FILL)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"war_sentiment_{ts}.xlsx",
    )


if __name__ == "__main__":
    log.info("Initializing with cached data...")
    _init_with_cache()
    log.info("Starting server → http://127.0.0.1:5000")
    # Disable the auto-reloader when using background threads in debug mode.
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
